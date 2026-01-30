"""
Upload de arquivos Excel para importação de pedidos.
Estrutura esperada: planilha com abas "Polpa congelada - ..." e "Manteiga de manga - ...".
Suporta revert por batch_id para desfazer uma importação.
"""
import math
import uuid
from io import BytesIO
from datetime import datetime
from typing import Optional, Any

from fastapi import APIRouter, File, UploadFile, HTTPException, Body
from openpyxl import load_workbook

from app.db import fatos, polpa, manteiga

router = APIRouter(prefix="/upload", tags=["upload"])

MESES_NOME = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def _safe_float(val: Any) -> Optional[float]:
    """Converte para float; aceita vírgula como decimal. Retorna None para vazio/NaN."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val):
            return None
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _safe_int(val: Any) -> Optional[int]:
    """Converte para int. Retorna None para vazio/NaN."""
    if val is None or val == "":
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        if math.isnan(val):
            return None
        return int(val) if val == int(val) else None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        try:
            return int(float(s.replace(",", ".")))
        except (ValueError, TypeError):
            return None
    return None


def _safe_str(val: Any) -> Optional[str]:
    """Retorna string ou None."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    s = str(val).strip()
    return s if s else None


def _parse_data_pedido(val: Any) -> Optional[datetime]:
    """Converte célula para datetime (data do pedido)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if hasattr(val, "date"):  # openpyxl datetime
        return val
    if isinstance(val, str):
        s = val.strip()[:10]
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except ValueError:
            try:
                return datetime.strptime(s, "%d/%m/%Y")
            except ValueError:
                return None
    return None


def _header_to_col(wb_sheet) -> dict:
    """Lê a primeira linha como cabeçalho e retorna mapeamento nome -> índice (1-based)."""
    headers = {}
    for col_idx, cell in enumerate(wb_sheet[1], start=1):
        name = (cell.value or "").strip().lower().replace(" ", "_")
        if name:
            headers[name] = col_idx
    return headers


def _cell_value(ws, row: int, col: int):
    """Valor da célula (openpyxl usa índices 1-based)."""
    return ws.cell(row=row, column=col).value


@router.post("/excel")
async def upload_excel(file: UploadFile = File(..., description="Arquivo .xlsx com abas Polpa e Manteiga")):
    """
    Importa pedidos a partir de um Excel.
    Estrutura esperada:
    - Aba cujo nome contém "Polpa": colunas fatos + logistica_brl, desconto_brl, lote_id, indice_qualidade_1a10, perda_processamento_pct
    - Aba cujo nome contém "Manteiga": colunas fatos + teor_umidade_pct, indice_acidez_mgKOH_g, ponto_fusao_c, indice_oxidacao_1a10, certificacao_exigida
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Envie um arquivo .xlsx")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Arquivo Excel inválido: {e}")

    batch_id = str(uuid.uuid4())
    inseridos_fatos = 0
    inseridos_polpa = 0
    inseridos_manteiga = 0
    erros = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Detectar tipo pela aba
        nome_lower = sheet_name.lower()
        if "polpa" in nome_lower:
            tipo_produto = "Polpa congelada"  # ou extrair do nome, ex: "Polpa congelada - Jan"
            is_polpa = True
        elif "manteiga" in nome_lower:
            tipo_produto = "Manteiga de manga"
            is_polpa = False
        else:
            continue

        headers = _header_to_col(ws)
        if not headers:
            erros.append(f"Aba '{sheet_name}': cabeçalho não encontrado.")
            continue

        # Mapear nomes possíveis (com e sem underscore, variações)
        def col(name_aliases: list) -> Optional[int]:
            for n in name_aliases:
                for k, idx in headers.items():
                    if n in k or k in n:
                        return idx
            return None

        idx_data = col(["data_pedido", "data"])
        idx_canal = col(["canal"])
        idx_regiao = col(["regiao_destino", "regiao"])
        idx_segmento = col(["cliente_segmento", "cliente_segmento", "segmento"])
        idx_qtde = col(["quantidade_kg", "quantidade"])
        idx_preco = col(["preco_unitario_brl_kg", "preco_unitario", "preco"])
        idx_nps = col(["nps_0a10", "nps"])

        if not idx_data or not idx_qtde or not idx_preco:
            erros.append(f"Aba '{sheet_name}': faltam colunas obrigatórias (data_pedido, quantidade_kg, preco_unitario_brl_kg).")
            continue

        list_fatos = []
        list_polpa = []
        list_manteiga = []

        for row_num in range(2, ws.max_row + 1):
            data_pedido = _parse_data_pedido(_cell_value(ws, row_num, idx_data))
            if not data_pedido:
                continue

            ano = data_pedido.year
            mes = data_pedido.month
            mes_nome = MESES_NOME.get(mes, "")
            mes_num = mes

            quantidade_kg = _safe_float(_cell_value(ws, row_num, idx_qtde))
            preco_kg = _safe_float(_cell_value(ws, row_num, idx_preco))
            if quantidade_kg is None and preco_kg is None:
                continue

            idx_row = row_num - 2
            id_pedido = f"{tipo_produto}_{ano}-{mes:02d}_{idx_row}"

            canal = _safe_str(_cell_value(ws, row_num, idx_canal)) if idx_canal else None
            regiao = _safe_str(_cell_value(ws, row_num, idx_regiao)) if idx_regiao else None
            cliente_segmento = _safe_str(_cell_value(ws, row_num, idx_segmento)) if idx_segmento else None
            nps = _safe_int(_cell_value(ws, row_num, idx_nps)) if idx_nps else None

            doc_fatos = {
                "id_pedido": id_pedido,
                "data_pedido": data_pedido,
                "tipo_produto": tipo_produto,
                "mes_do_ano": mes_nome,
                "mes_do_ano_num": mes_num,
                "canal": canal,
                "regiao_destino": regiao,
                "cliente_segmento": cliente_segmento,
                "quantidade_kg": quantidade_kg,
                "preco_unitario_brl_kg": preco_kg,
                "nps_0a10": nps,
                "import_batch_id": batch_id,
            }
            list_fatos.append(doc_fatos)

            if is_polpa:
                idx_log = col(["logistica_brl", "logistica"])
                idx_desconto = col(["desconto_brl", "desconto"])
                idx_lote = col(["lote_id", "lote"])
                idx_qualidade = col(["indice_qualidade_1a10", "indice_qualidade", "qualidade"])
                idx_perda = col(["perda_processamento_pct", "perda_processamento", "perda"])
                list_polpa.append({
                    "id_pedido": id_pedido,
                    "logistica_brl": _safe_float(_cell_value(ws, row_num, idx_log)) if idx_log else None,
                    "desconto_brl": _safe_float(_cell_value(ws, row_num, idx_desconto)) if idx_desconto else None,
                    "lote_id": _safe_str(_cell_value(ws, row_num, idx_lote)) if idx_lote else None,
                    "indice_qualidade_1a10": _safe_int(_cell_value(ws, row_num, idx_qualidade)) if idx_qualidade else None,
                    "perda_processamento_pct": _safe_float(_cell_value(ws, row_num, idx_perda)) if idx_perda else None,
                    "import_batch_id": batch_id,
                })
            else:
                idx_umidade = col(["teor_umidade_pct", "teor_umidade", "umidade"])
                idx_acidez = col(["indice_acidez_mgKOH_g", "indice_acidez", "acidez"])
                idx_fusao = col(["ponto_fusao_c", "ponto_fusao", "fusao"])
                idx_oxidacao = col(["indice_oxidacao_1a10", "indice_oxidacao", "oxidacao"])
                idx_cert = col(["certificacao_exigida", "certificacao"])
                list_manteiga.append({
                    "id_pedido": id_pedido,
                    "teor_umidade_pct": _safe_float(_cell_value(ws, row_num, idx_umidade)) if idx_umidade else None,
                    "indice_acidez_mgKOH_g": _safe_float(_cell_value(ws, row_num, idx_acidez)) if idx_acidez else None,
                    "ponto_fusao_c": _safe_float(_cell_value(ws, row_num, idx_fusao)) if idx_fusao else None,
                    "indice_oxidacao_1a10": _safe_int(_cell_value(ws, row_num, idx_oxidacao)) if idx_oxidacao else None,
                    "certificacao_exigida": _safe_str(_cell_value(ws, row_num, idx_cert)) if idx_cert else None,
                    "import_batch_id": batch_id,
                })

        if list_fatos:
            await fatos.insert_many(list_fatos)
            inseridos_fatos += len(list_fatos)
        if list_polpa:
            await polpa.insert_many(list_polpa)
            inseridos_polpa += len(list_polpa)
        if list_manteiga:
            await manteiga.insert_many(list_manteiga)
            inseridos_manteiga += len(list_manteiga)

    return {
        "message": "Importação concluída.",
        "batch_id": batch_id,
        "inseridos": {
            "fatos_pedidos": inseridos_fatos,
            "polpa_metricas": inseridos_polpa,
            "manteiga_metricas": inseridos_manteiga,
        },
        "erros": erros if erros else None,
    }


@router.post("/revert")
async def revert_import(
    batch_id: str = Body(..., embed=True, description="batch_id retornado pelo POST /upload/excel"),
):
    """
    Remove todos os documentos inseridos em uma importação (identificada por batch_id).
    Use o batch_id retornado na resposta do POST /upload/excel.
    """
    if not batch_id or not batch_id.strip():
        raise HTTPException(400, "batch_id é obrigatório.")

    result_fatos = await fatos.delete_many({"import_batch_id": batch_id})
    result_polpa = await polpa.delete_many({"import_batch_id": batch_id})
    result_manteiga = await manteiga.delete_many({"import_batch_id": batch_id})

    return {
        "message": "Importação revertida.",
        "batch_id": batch_id,
        "removidos": {
            "fatos_pedidos": result_fatos.deleted_count,
            "polpa_metricas": result_polpa.deleted_count,
            "manteiga_metricas": result_manteiga.deleted_count,
        },
    }
